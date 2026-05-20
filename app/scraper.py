import requests
import cloudscraper
import re
from datetime import datetime
from io import BytesIO
from pypdf import PdfReader
from openpyxl import load_workbook
import xlrd
from config.settings import REQUEST_TIMEOUT

QUARTER_MAP = {
    "Q1": "tw1",
    "Q2": "tw2",
    "Q3": "tw3",
    "Q4": "tw4",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": "https://www.idx.co.id/",
}

BASE_URL = "https://www.idx.co.id"
SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls"}
MAX_ATTACHMENTS_TO_PARSE = 4
MAX_CHARS_PER_FILE = 8000
MAX_TOTAL_CHARS = 14000

SHARES_ANNOUNCEMENT_KEYWORD = "Laporan Bulanan Registrasi Pemegang Efek"
SHARES_OUTSTANDING_KEYWORDS = [
    "shares outstanding",
    "total shares outstanding",
    "saham beredar",
    "jumlah saham beredar",
    "jumlah lembar saham beredar",
    "outstanding shares",
]
SHARES_FLOAT_KEYWORDS = [
    "shares float",
    "free float",
    "saham float",
    "saham publik",
    "float shares",
]
SHARES_INSTITUTIONAL_KEYWORDS = [
    "shares institutional",
    "institutional",
    "kepemilikan institusional",
    "saham institusional",
]
SHARES_INSIDER_KEYWORDS = [
    "shares insider",
    "insider",
    "kepemilikan insider",
    "saham insider",
    "manajemen dan insider",
]

PRIORITY_KEYWORDS = [
    "financial",
    "keuangan",
    "laporan-keuangan",
    "quarter",
    "kuartal",
    "interim",
]
LOW_PRIORITY_KEYWORDS = [
    "esg",
    "sustainability",
    "keberlanjutan",
]

FINANCIAL_TEXT_KEYWORDS = [
    "jumlah aset",
    "total assets",
    "jumlah liabilitas",
    "total liabilities",
    "jumlah ekuitas",
    "total equity",
    "pendapatan",
    "revenue",
    "laba operasional",
    "operating profit",
    "laba bersih",
    "net profit",
    "laba tahun berjalan",
    "profit for the period",
    "beban operasional",
    "operating expense",
    "net interest income",
    "current ratio",
    "debt to equity",
    "return on assets",
    "return on equity",
    "eps",
    "book value per share",
]

# Shareholder keywords — nama & angka sering di baris terpisah, perlu context window lebih lebar
SHAREHOLDER_TEXT_KEYWORDS = [
    "pemegang saham",
    "shareholder",
    "komposisi pemegang saham",
    "composition of shareholders",
    "jumlah lembar saham",
    "number of shares",
    "persentase kepemilikan",
    "percentage of ownership",
    "kepemilikan saham",
    "share ownership",
    "daftar pemegang",
]

NUMERIC_ROW_RE = re.compile(r"\d[\d,\.]*")


def _create_scraper():
    return cloudscraper.create_scraper(
        browser={"browser": "chrome", "platform": "windows", "mobile": False}
    )

def _get(url: str, params: dict) -> dict:
    """Helper function to make GET requests to IDX API"""
    try:
        # IDX often blocks plain requests clients; cloudscraper handles anti-bot checks.
        scraper = _create_scraper()
        response = scraper.get(
            url, params=params, headers=HEADERS, timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()
        return response.json()
    except requests.HTTPError as exc:
        raise RuntimeError(
            f"IDX API returned HTTP {exc.response.status_code} for {url}"
        ) from exc
    except ValueError:
        # Fallback to plain requests when response body is not valid JSON from scraper.
        response = requests.get(
            url, params=params, headers=HEADERS, timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()
        return response.json()
    except requests.RequestException as exc:
        raise RuntimeError(f"Request to IDX API failed: {exc}") from exc


def _download_file(url: str) -> bytes:
    try:
        scraper = _create_scraper()
        response = scraper.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.content
    except requests.RequestException as exc:
        raise RuntimeError(f"Failed to download IDX attachment: {exc}") from exc


def _extract_pdf_text(content: bytes) -> str:
    reader = PdfReader(BytesIO(content))
    text_parts = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            text_parts.append(page_text)
    return "\n".join(text_parts).strip()


def _extract_xlsx_text(content: bytes) -> str:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        rows.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            clean_cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if clean_cells:
                rows.append(" | ".join(clean_cells))
    wb.close()
    return "\n".join(rows).strip()


def _extract_xls_text(content: bytes) -> str:
    wb = xlrd.open_workbook(file_contents=content)
    rows = []
    for sheet in wb.sheets():
        rows.append(f"[Sheet: {sheet.name}]")
        for i in range(sheet.nrows):
            values = sheet.row_values(i)
            clean_cells = [str(cell).strip() for cell in values if str(cell).strip()]
            if clean_cells:
                rows.append(" | ".join(clean_cells))
    return "\n".join(rows).strip()


def _extract_attachment_text(file_name: str, content: bytes) -> str:
    lower_name = file_name.lower()
    if lower_name.endswith(".pdf"):
        return _extract_pdf_text(content)
    if lower_name.endswith(".xlsx"):
        return _extract_xlsx_text(content)
    if lower_name.endswith(".xls"):
        return _extract_xls_text(content)
    return ""


def _focus_financial_text(raw_text: str) -> str:
    lines = [line.strip() for line in (raw_text or "").splitlines() if line.strip()]
    if not lines:
        return ""

    selected_indexes = set()
    for i, line in enumerate(lines):
        lower_line = line.lower()

        # Financial metrics: harus ada keyword DAN angka di baris yang sama
        has_metric_keyword = any(keyword in lower_line for keyword in FINANCIAL_TEXT_KEYWORDS)
        has_numeric_value = bool(NUMERIC_ROW_RE.search(line))
        if has_metric_keyword and has_numeric_value:
            selected_indexes.add(i)
            for offset in (-1, 1):
                j = i + offset
                if 0 <= j < len(lines):
                    selected_indexes.add(j)

        # Shareholder: cukup ada keyword saja, ambil konteks +-8 baris
        # karena nama pemegang saham dan angka sering di baris yang berbeda
        has_shareholder_keyword = any(keyword in lower_line for keyword in SHAREHOLDER_TEXT_KEYWORDS)
        if has_shareholder_keyword:
            for offset in range(-8, 9):
                j = i + offset
                if 0 <= j < len(lines):
                    selected_indexes.add(j)

    if not selected_indexes:
        return "\n".join(lines[:140])

    focused_lines = [lines[i] for i in sorted(selected_indexes)]
    return "\n".join(focused_lines[:300])


def _parse_idx_datetime(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    candidate_formats = [
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%d/%m/%Y %I:%M:%S %p",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%Y%m%d",
    ]

    iso_candidates = [text]
    if text.endswith("Z"):
        iso_candidates.append(text[:-1] + "+00:00")

    for candidate in iso_candidates:
        try:
            return datetime.fromisoformat(candidate)
        except ValueError:
            continue

    for candidate_format in candidate_formats:
        try:
            parsed = datetime.strptime(text, candidate_format)
            if parsed.tzinfo is not None:
                parsed = parsed.replace(tzinfo=None)
            return parsed
        except ValueError:
            continue

    return None


def _format_report_date(value):
    parsed = _parse_idx_datetime(value)
    if not parsed:
        return None
    return parsed.date().isoformat()


def _normalize_announcement_reply(reply: dict) -> dict:
    pengumuman = reply.get("pengumuman") or {}
    attachments = reply.get("attachments") or reply.get("Attachments") or []
    title = str(pengumuman.get("JudulPengumuman") or "").strip()
    announcement_dt = _parse_idx_datetime(pengumuman.get("TglPengumuman") or pengumuman.get("CreatedDate"))

    return {
        "title": title,
        "date": announcement_dt,
        "month_key": announcement_dt.strftime("%Y-%m") if announcement_dt else None,
        "attachments": attachments,
        "raw": reply,
    }


def _select_monthly_announcements(replies: list[dict]) -> list[dict]:
    grouped = {}
    for reply in replies:
        normalized = _normalize_announcement_reply(reply)
        month_key = normalized.get("month_key")
        if not month_key:
            continue

        current = grouped.get(month_key)
        if current is None:
            grouped[month_key] = normalized
            continue

        current_title = str(current.get("title") or "").lower()
        candidate_title = str(normalized.get("title") or "").lower()
        current_is_correction = "koreksi" in current_title
        candidate_is_correction = "koreksi" in candidate_title

        if candidate_is_correction and not current_is_correction:
            grouped[month_key] = normalized
            continue

        if candidate_is_correction == current_is_correction:
            current_date = current.get("date")
            candidate_date = normalized.get("date")
            if candidate_date and (not current_date or candidate_date > current_date):
                grouped[month_key] = normalized

    return sorted(
        grouped.values(),
        key=lambda item: item.get("date") or datetime.min,
    )


def _clean_numeric_value(value):
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    normalized = re.sub(r"[^\d]", "", text)
    if len(normalized) < 4:
        return None

    try:
        number = int(normalized)
    except ValueError:
        return None

    return number if number > 0 else None


def _normalize_share_report_text(text: str) -> str:
    normalized = str(text or "")
    normalized = re.sub(r"(?<=[.,])\s+(?=\d)", "", normalized)
    normalized = re.sub(r"(?<=\d)\s+(?=[.,]\d)", "", normalized)
    return normalized


def _extract_integer_values(text: str, include_zero: bool = False) -> list[int]:
    values = []
    for match in re.finditer(r"-|\d{1,3}(?:[.,]\d{3})+|\d+", text or ""):
        token = match.group(0)
        tail = text[match.end():match.end() + 3]
        if "%" in tail:
            continue
        if token == "-":
            if include_zero:
                values.append(0)
            continue

        cleaned = _clean_numeric_value(token)
        if cleaned is not None:
            values.append(cleaned)
        elif include_zero and re.fullmatch(r"0+", token):
            values.append(0)
    return values


def _extract_values_after_phrases(
    text: str,
    phrases: list[str],
    window: int = 400,
    include_zero: bool = False,
) -> list[int]:
    normalized = _normalize_share_report_text(text)
    lower_text = normalized.lower()

    for phrase in phrases:
        index = lower_text.find(phrase.lower())
        if index == -1:
            continue
        segment = normalized[index + len(phrase):index + len(phrase) + window]
        values = _extract_integer_values(segment, include_zero=include_zero)
        if values:
            return values

    return []


def _extract_first_after_phrases(
    text: str,
    phrases: list[str],
    window: int = 400,
    include_zero: bool = False,
) -> int | None:
    values = _extract_values_after_phrases(
        text,
        phrases,
        window=window,
        include_zero=include_zero,
    )
    return values[0] if values else None


def _extract_best_first_after_phrases(
    text: str,
    phrases: list[str],
    window: int = 400,
    include_zero: bool = False,
) -> int | None:
    normalized = _normalize_share_report_text(text)
    lower_text = normalized.lower()
    first_values = []

    for phrase in phrases:
        lower_phrase = phrase.lower()
        start = 0
        while True:
            index = lower_text.find(lower_phrase, start)
            if index == -1:
                break
            segment = normalized[index + len(phrase):index + len(phrase) + window]
            values = _extract_integer_values(segment, include_zero=include_zero)
            if values:
                first_values.append(values[0])
            start = index + len(phrase)

    return max(first_values) if first_values else None


def _extract_max_after_phrases(text: str, phrases: list[str], window: int = 400) -> int | None:
    values = _extract_values_after_phrases(text, phrases, window=window)
    return max(values) if values else None


def _extract_free_float_table_values(text: str) -> list[int]:
    return _extract_values_after_phrases(
        text,
        [
            "Jumlah saham Free Float",
            "The amount of Free Float Share",
        ],
        window=600,
        include_zero=True,
    )


def _extract_reported_free_float(text: str) -> int | None:
    values = _extract_free_float_table_values(text)
    if values:
        return values[0]

    values = _extract_values_after_phrases(
        text,
        [
            "Informasi Saham Free Float",
            "Free Float Share Information",
        ],
        window=500,
        include_zero=True,
    )
    if len(values) >= 2:
        return values[1]
    if values:
        return values[0]
    return None


def _extract_bod_boc_shares(text: str) -> int | None:
    normalized = _normalize_share_report_text(text)
    section_patterns = [
        (
            "Investor Type & Classification Number of Shares",
            "Number of Scripless Shares based on Investor Type and Classification from KSEI",
        ),
        (
            "Tipe dan Klasifikasi Investor Jumlah Saham",
            "Jumlah Saham Scripless berdasarkan Tipe dan Klasifikasi Investor dari KSEI",
        ),
    ]

    for start_marker, end_marker in section_patterns:
        start = normalized.find(start_marker)
        if start == -1:
            continue
        end = normalized.find(end_marker, start)
        if end == -1:
            continue

        section = normalized[start:end]
        if "BOD / BOC" not in section and "Direksi dan Dewan Komisaris" not in section:
            continue

        values = _extract_integer_values(section, include_zero=True)
        if len(values) >= 2:
            return values[1]

    return None


def _extract_director_commissioner_ownership(text: str) -> int | None:
    normalized = _normalize_share_report_text(text)
    section_markers = [
        "Laporan Kepemilikan Saham Oleh Direksi dan Komisaris",
        "Share Ownership Report by Directors and Commissioners",
    ]
    end_markers = [
        "Informasi Saham Free Float",
        "Free Float Share Information",
    ]

    for marker in section_markers:
        start = normalized.find(marker)
        if start == -1:
            continue

        end_candidates = [
            normalized.find(end_marker, start)
            for end_marker in end_markers
            if normalized.find(end_marker, start) != -1
        ]
        end = min(end_candidates) if end_candidates else start + 5000
        section = normalized[start:end]

        current_values = []
        for match in re.finditer(
            r"(\d{1,3}(?:[.,]\d{3})+)\s+\d+(?:[,.]\d+)?%\s+"
            r"(\d{1,3}(?:[.,]\d{3})+)\s+\d+(?:[,.]\d+)?%",
            section,
        ):
            cleaned = _clean_numeric_value(match.group(2))
            if cleaned is not None:
                current_values.append(cleaned)

        if current_values:
            return sum(current_values)

    return None


def _extract_free_float_bod_boc_shares(text: str) -> int | None:
    values = _extract_free_float_table_values(text)
    unique_values = []
    for value in values:
        if value not in unique_values:
            unique_values.append(value)

    if len(unique_values) >= 2:
        candidate = unique_values[1]
        if candidate < 1_000_000_000:
            return candidate
    return None


def _extract_shares_metrics_from_report(text: str) -> dict:
    shares_outstanding = _extract_max_after_phrases(
        text,
        [
            "Total",
        ],
        window=160,
    )
    shares_float = _extract_reported_free_float(text)
    shares_institutional = _extract_best_first_after_phrases(
        text,
        [
            "Financial Institutional (IB)",
        ],
        window=80,
        include_zero=True,
    )
    shares_insider = (
        _extract_bod_boc_shares(text)
        or _extract_director_commissioner_ownership(text)
        or _extract_free_float_bod_boc_shares(text)
    )

    return {
        "sharesOutstanding": shares_outstanding,
        "sharesFloat": shares_float,
        "sharesInstitutional": shares_institutional,
        "sharesInsider": shares_insider,
    }


def _extract_metric_from_text(lines: list[str], keywords: list[str]) -> int | None:
    lowered_lines = [line.lower() for line in lines]

    for index, lower_line in enumerate(lowered_lines):
        if not any(keyword in lower_line for keyword in keywords):
            continue

        same_line_matches = []
        for match in re.finditer(r"\d{1,3}(?:[.,]\d{3})+|\d+", lines[index]):
            cleaned = _clean_numeric_value(match.group(0))
            if cleaned is not None:
                same_line_matches.append(cleaned)
        if same_line_matches:
            return max(same_line_matches)

        search_window = "\n".join(lines[index:index + 4])
        numeric_matches = []
        for match in re.finditer(r"\d{1,3}(?:[.,]\d{3})+|\d+", search_window):
            cleaned = _clean_numeric_value(match.group(0))
            if cleaned is not None:
                numeric_matches.append(cleaned)

        if numeric_matches:
            return max(numeric_matches)

    return None


def _parse_shares_report_text(text: str, announcement_date: str | None = None) -> dict:
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    report_metrics = _extract_shares_metrics_from_report(text)
    return {
        "date": announcement_date,
        "sharesOutstanding": report_metrics.get("sharesOutstanding") or _extract_metric_from_text(lines, SHARES_OUTSTANDING_KEYWORDS),
        "sharesFloat": report_metrics.get("sharesFloat") or _extract_metric_from_text(lines, SHARES_FLOAT_KEYWORDS),
        "sharesInstitutional": report_metrics.get("sharesInstitutional") or _extract_metric_from_text(lines, SHARES_INSTITUTIONAL_KEYWORDS),
        "sharesInsider": report_metrics.get("sharesInsider") or _extract_metric_from_text(lines, SHARES_INSIDER_KEYWORDS),
    }


def _fill_missing_share_metrics(items: list[dict]) -> list[dict]:
    carry_fields = ["sharesOutstanding", "sharesFloat", "sharesInsider"]

    for field in carry_fields:
        last_value = None
        for item in items:
            if item.get(field) is None:
                item[field] = last_value
            else:
                last_value = item.get(field)

        next_value = None
        for item in reversed(items):
            if item.get(field) is None:
                item[field] = next_value
            else:
                next_value = item.get(field)

    for item in items:
        if item.get("sharesInstitutional") is None:
            item["sharesInstitutional"] = 0

    return items


def _get_announcement_full_save_path(reply: dict) -> str:
    attachments = reply.get("attachments") or reply.get("Attachments") or []
    for attachment in attachments:
        full_save_path = str(attachment.get("FullSavePath") or attachment.get("File_Path") or "").strip()
        if full_save_path.lower().endswith(".pdf"):
            return full_save_path
    first_attachment = attachments[0] if attachments else {}
    return str(first_attachment.get("FullSavePath") or first_attachment.get("File_Path") or "").strip()


def fetch_idx_shares_announcements(symbol: str, date_from: str = "19010101", date_to: str | None = None) -> list[dict]:
    url = "https://www.idx.co.id/primary/ListedCompany/GetAnnouncement"
    params = {
        "kodeEmiten": symbol.upper(),
        "emitenType": "*",
        "indexFrom": 0,
        "pageSize": 100,
        "dateFrom": date_from,
        "dateTo": date_to or datetime.now().strftime("%Y%m%d"),
        "lang": "id",
        "keyword": SHARES_ANNOUNCEMENT_KEYWORD,
    }
    response = _get(url, params)
    replies = response.get("Replies") or response.get("Results") or []
    return replies if isinstance(replies, list) else []


def scrape_shares_data(symbol: str) -> dict:
    replies = fetch_idx_shares_announcements(symbol)
    selected_replies = _select_monthly_announcements(replies)

    items = []
    for reply in selected_replies:
        raw_reply = reply.get("raw") or {}
        pengumuman = raw_reply.get("pengumuman") or {}
        announcement_date = _format_report_date(pengumuman.get("TglPengumuman") or pengumuman.get("CreatedDate") or reply.get("date"))
        file_url = _get_announcement_full_save_path(raw_reply or reply)
        if not file_url:
            continue

        try:
            content = _download_file(file_url)
            extracted_text = _extract_pdf_text(content)
        except Exception:
            extracted_text = ""

        item = _parse_shares_report_text(extracted_text, announcement_date=announcement_date)
        if item.get("date") is None:
            item["date"] = announcement_date
        items.append(item)

    items.sort(key=lambda item: item.get("date") or "")
    _fill_missing_share_metrics(items)

    return {
        "symbol": symbol.upper(),
        "count": len(items),
        "items": items,
    }

def _collect_report_text(raw_data: dict) -> tuple[str, list[dict]]:
    attachments = raw_data.get("Attachments") or []
    eligible = []

    for item in attachments:
        file_name = str(item.get("File_Name") or "")
        ext = str(item.get("File_Type") or "").lower()
        if not ext and "." in file_name:
            ext = "." + file_name.lower().split(".")[-1]
        if ext in SUPPORTED_EXTENSIONS:
            eligible.append(item)

    def score_attachment(item: dict) -> int:
        file_name = str(item.get("File_Name") or "").lower()
        score = 0
        for kw in PRIORITY_KEYWORDS:
            if kw in file_name:
                score += 3
        for kw in LOW_PRIORITY_KEYWORDS:
            if kw in file_name:
                score -= 2
        return score

    eligible.sort(key=score_attachment, reverse=True)

    parsed_docs = []
    text_chunks = []
    total_chars = 0

    for item in eligible[:MAX_ATTACHMENTS_TO_PARSE]:
        file_name = str(item.get("File_Name") or "unknown")
        file_path = str(item.get("File_Path") or "")
        file_url = file_path if file_path.startswith("http") else f"{BASE_URL}{file_path}"

        try:
            content = _download_file(file_url)
            extracted = _extract_attachment_text(file_name, content)
            extracted = _focus_financial_text(extracted)
        except Exception:
            extracted = ""

        extracted = extracted.strip()
        extracted = extracted[:MAX_CHARS_PER_FILE]

        parsed_docs.append(
            {
                "file_name": file_name,
                "file_type": item.get("File_Type"),
                "file_url": file_url,
                "extracted_chars": len(extracted),
            }
        )

        if extracted:
            remaining = MAX_TOTAL_CHARS - total_chars
            if remaining <= 0:
                break
            excerpt = extracted[:remaining]
            total_chars += len(excerpt)
            text_chunks.append(f"### Dokumen: {file_name}\n{excerpt}")

    return "\n\n".join(text_chunks).strip(), parsed_docs

def _normalized_lookup(raw_data: dict) -> dict:
    lookup = {}
    for key, value in (raw_data or {}).items():
        normalized = "".join(ch for ch in str(key).lower() if ch.isalnum())
        if normalized and normalized not in lookup:
            lookup[normalized] = value
    return lookup


def _pick_field(raw_data: dict, *aliases: str):
    for alias in aliases:
        if alias in raw_data and raw_data.get(alias) not in (None, ""):
            return raw_data.get(alias)

    normalized_lookup = _normalized_lookup(raw_data)
    for alias in aliases:
        normalized = "".join(ch for ch in str(alias).lower() if ch.isalnum())
        value = normalized_lookup.get(normalized)
        if value not in (None, ""):
            return value

    return None


def _fetch_report_results(symbol: str, year: int, periode: str, report_type: str = "rdf") -> list[dict]:
    url = "https://www.idx.co.id/primary/ListedCompany/GetFinancialReport"
    params = {
        "ReportType": report_type,
        "KodeEmiten": symbol.upper(),
        "Year": str(year),
        "SortColumn": "KodeEmiten",
        "SortOrder": "asc",
        "EmitenType": "s",
        "Periode": periode,
        "indexfrom": 1,
        "pagesize": 12,
    }

    response_data = _get(url, params)
    return response_data.get("Results") or []


def get_financial_report(symbol: str, year: int, quarter: str | None = None) -> dict:
    """
    Get financial report from IDX API
    
    Args:
        symbol: Stock code (e.g., 'BBRI', 'ASII')
        year: Year of report (e.g., 2024)
        quarter: Quarter (Q1, Q2, Q3, Q4), optional for yearly mode
    
    Returns:
        dict: Financial report data
    """
    requested_quarter = (quarter or "").strip().upper()
    periode = QUARTER_MAP.get(requested_quarter) if requested_quarter else "audit"

    try:
        # Primary mode from IDX search contract: reportType=rdf and periode by request.
        results = _fetch_report_results(symbol, year, periode, report_type="rdf")
        if results:
            return results[0]

        # Fallback for potential period casing differences.
        fallback_results = _fetch_report_results(symbol, year, periode.upper(), report_type="rdf")
        if fallback_results:
            return fallback_results[0]

        # Compatibility fallback.
        legacy_results = _fetch_report_results(symbol, year, periode, report_type="PDF")
        if legacy_results:
            return legacy_results[0]

        raise RuntimeError(f"No financial data found for {symbol} - {year} {periode}")
            
    except Exception as exc:
        raise RuntimeError(f"Failed to fetch financial report: {exc}") from exc

def parse_financial_data(raw_data: dict) -> dict:
    """Parse raw IDX API response into standardized format"""
    return {
        "kode_emiten": _pick_field(raw_data, "KodeEmiten", "Code"),
        "nama_emiten": _pick_field(raw_data, "NamaEmiten", "Name"),
        "periode_laporan": _pick_field(raw_data, "PeriodeLaporan", "Report_Period", "Period"),
        "tanggal_laporan": _pick_field(raw_data, "TanggalLaporan", "Report_Date", "File_Modified", "Date"),
        "sector": _pick_field(raw_data, "Sector", "Sektor"),
        "sub_sector": _pick_field(raw_data, "SubSector", "Sub_Sector", "SubSektor"),
        "revenue": _pick_field(raw_data, "Revenue", "TotalRevenue", "Sales"),
        "cost_of_goods_sold": _pick_field(raw_data, "CostOfGoodsSold", "COGS"),
        "gross_profit": _pick_field(raw_data, "GrossProfit"),
        "operating_expense": _pick_field(raw_data, "OperatingExpense", "OperatingExpenses"),
        "operating_profit": _pick_field(raw_data, "OperatingProfit", "OperatingIncome"),
        "net_profit": _pick_field(raw_data, "NetProfit", "NetIncome", "ProfitForTheYear", "ProfitLoss"),
        "total_assets": _pick_field(raw_data, "TotalAssets", "TotalAsset", "Assets"),
        "total_liabilities": _pick_field(raw_data, "TotalLiabilities", "Liabilities", "TotalLiability"),
        "total_equity": _pick_field(raw_data, "TotalEquity", "Equity"),
        "eps": _pick_field(raw_data, "EPS", "EarningPerShare"),
        "book_value_per_share": _pick_field(raw_data, "BookValuePerShare", "BVPS"),
        "roe": _pick_field(raw_data, "ROE", "ReturnOnEquity"),
        "roa": _pick_field(raw_data, "ROA", "ReturnOnAssets"),
        "npm": _pick_field(raw_data, "NPM", "NetMargin", "NetProfitMargin"),
        "der": _pick_field(raw_data, "DER", "DebtToEquity"),
        "per": _pick_field(raw_data, "PER", "PriceEarningsRatio"),
        "pbr": _pick_field(raw_data, "PBR", "PriceToBookRatio"),
        "current_ratio": _pick_field(raw_data, "CurrentRatio"),
    }

def scrape_fundamental(symbol: str, year: int, quarter: str | None = None) -> dict:
    """Main function to scrape fundamental data from IDX"""
    raw_data = get_financial_report(symbol, year, quarter)

    parsed_data = parse_financial_data(raw_data)
    report_text, parsed_documents = _collect_report_text(raw_data)
    request_period = (quarter or "").strip().upper() or "AUDIT"

    return {
        "symbol": symbol.upper(),
        "year": year,
        "quarter": request_period,
        "data": parsed_data,
        "report_text": report_text,
        "report_documents": parsed_documents,
        "raw_response": raw_data,
    }

def find_shareholders(data: str) -> list[dict]:
    """
    Extract all shareholders from report text.
    Returns a list of dicts with keys: name, shares, ownership.
    Returns empty list if none found.
    """
    import re

    if not data:
        return []

    print("data length:", len(data))
    text = data

    # Cari semua angka besar (format ribuan dengan titik, misal 1.234.567.890)
    matches = list(re.finditer(r"\d{1,3}(?:\.\d{3}){2,}", text))

    if not matches:
        return []

    print(f"Found {len(matches)} numeric candidates in text.")
    candidates = []
    seen_shares = set()

    for m in matches:
        raw_number = m.group(0)
        shares = int(raw_number.replace(".", ""))

        # Skip angka terlalu kecil (bukan jumlah saham signifikan)
        if shares < 1_000_000:
            continue

        # Deduplicate angka yang sama persis
        if shares in seen_shares:
            continue
        seen_shares.add(shares)

        start = max(0, m.start() - 150)
        end = min(len(text), m.end() + 150)
        context = text[start:end]

        # Cari persentase kepemilikan di sekitar angka
        percent_match = re.search(r"\b\d{1,3}[,\.]\d{2}\b", context)
        ownership = None
        if percent_match:
            try:
                ownership = float(percent_match.group(0).replace(",", "."))
                # Validasi range persentase wajar
                if not (0 < ownership <= 100):
                    ownership = None
            except ValueError:
                ownership = None

        # Ambil nama dari baris sebelum angka
        before_number = context.split(raw_number)[0]
        lines = [l.strip() for l in before_number.strip().split("\n") if l.strip()]
        name = lines[-1] if lines else ""
        name = re.sub(r"\d[\d\.,]*", "", name).strip()
        name = re.sub(r"\s{2,}", " ", name).strip(" |:-")

        if not name:
            continue

        #  FILTER WAJIB: harus ada konteks shareholder
        if not any(k in context.lower() for k in [
            "pemegang saham",
            "shareholder",
            "komposisi pemegang saham"
        ]):
            continue

        # FILTER NOISE (hindari modal/penerbitan)
        if any(bad in context.lower() for bad in [
            "modal",
            "penerbitan",
            "issued",
            "capital",
            "treasury"
        ]):
            continue

        candidates.append({
            "name": name,
            "shares": shares,
            "ownership": ownership,
        })

    # Urutkan dari terbesar ke terkecil
    candidates.sort(key=lambda x: x["shares"], reverse=True)

    return candidates


# Alias untuk backward compatibility
def find_largest_shareholder(data: str):
    shareholders = find_shareholders(data)
    if not shareholders:
        print("No shareholders found in text.")
        return None 
    return shareholders[0]
